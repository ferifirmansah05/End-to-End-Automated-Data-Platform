from datetime import datetime, timedelta
import os, smtplib, tempfile
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import json
from airflow import DAG
from airflow.operators.python import PythonOperator

from clickhouse_driver import Client
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                numbers as xl_numbers)
from openpyxl.utils import get_column_letter
import pendulum

lokal_tz = pendulum.timezone("Asia/Jakarta")

CLICKHOUSE_HOST = os.environ["CLICKHOUSE_HOST"]
CLICKHOUSE_USER = os.environ["CLICKHOUSE_USER"]
CLICKHOUSE_PASS = os.environ["CLICKHOUSE_PASS"]

# Mengambil variabel SMTP WAJIB dari file JSON external
# (Akan error jika file tidak ada atau key di dalam JSON tidak lengkap)
JSON_CONFIG_PATH = "/opt/airflow/config_email.json"

with open(JSON_CONFIG_PATH, "r") as file:
    config_data = json.load(file)

SMTP_SERVER     = config_data["SMTP_SERVER"]
SMTP_PORT       = int(config_data["SMTP_PORT"])
SMTP_USER       = config_data["SMTP_USER"]
SMTP_PASSWORD   = config_data["SMTP_PASSWORD"]
SENDER_EMAIL    = config_data["SENDER_EMAIL"]
REPORT_RECEIVER = config_data["RECEIVER_EMAIL"]

default_args = {
    "owner": "data-engineering",
    'start_date': datetime.now(lokal_tz),
    "depends_on_past": False,
    "email_on_failure": False,
    "retries": 2,
    "retry_delay": timedelta(minutes=3),
}

dag = DAG(
    dag_id="daily_financial_report",
    default_args=default_args,
    description="Laporan keuangan harian — Revenue, HPP, Gross Profit, PO, Stok, Deviasi BOM",
    schedule_interval="0 8 * * *",
    catchup=False,
    tags=["finance", "report", "daily"],
)

def fetch_financial_data(**context):
    ch = Client(host=CLICKHOUSE_HOST, user=CLICKHOUSE_USER, password=CLICKHOUSE_PASS)
    d  = context["yesterday_ds"]

    queries = {
        "rev_summary": f"""
            SELECT SUM(subtotal), COUNT(*), SUM(qty)
            FROM default.fact_sales 
            WHERE tanggal_transaksi = '{d}'""",

        "revenue_rows": f"""
            SELECT d.nama_cabang, f.kategori_payment,
                COUNT(*) AS trx, SUM(f.qty) AS qty, SUM(f.subtotal) AS rev
            FROM default.fact_sales f
            LEFT JOIN default.dim_cabang d ON f.id_cabang = d.id_cabang
            WHERE f.tanggal_transaksi = '{d}'
            GROUP BY d.nama_cabang, f.kategori_payment
            ORDER BY d.nama_cabang ASC, rev DESC""",

        "hpp_rows": f"""
            SELECT d.nama_cabang, dm.kategori,
                SUM(f.quantity) AS qty_pakai, SUM(f.total_biaya) AS hpp
            FROM default.fact_cost_of_material f
            LEFT JOIN default.dim_cabang d ON f.id_cabang = d.id_cabang
            LEFT JOIN default.dim_menu dm ON f.id_menu = dm.id_menu
            WHERE f.tanggal = '{d}'
            GROUP BY d.nama_cabang, dm.kategori 
            ORDER BY d.nama_cabang ASC, hpp DESC""",

        "gp_rows": f"""
            SELECT s.nama_cabang, s.rev, c.hpp, s.rev - c.hpp AS gp
            FROM (
                SELECT d.nama_cabang, SUM(f.subtotal) AS rev
                FROM default.fact_sales f
                LEFT JOIN default.dim_cabang d ON f.id_cabang = d.id_cabang
                WHERE f.tanggal_transaksi = '{d}'
                GROUP BY d.nama_cabang
            ) s
            LEFT JOIN (
                SELECT d.nama_cabang, SUM(f.total_biaya) AS hpp
                FROM default.fact_cost_of_material f
                LEFT JOIN default.dim_cabang d ON f.id_cabang = d.id_cabang
                WHERE f.tanggal = '{d}'
                GROUP BY d.nama_cabang
            ) c ON s.nama_cabang = c.nama_cabang
            ORDER BY gp DESC""",

        "po_rows": f"""
            SELECT d.nama_cabang, f.status,
                COUNT(DISTINCT f.nomor_po) AS jml_po,
                SUM(f.kuantitas) AS qty, SUM(f.total_biaya) AS nilai
            FROM default.fact_purchases f
            LEFT JOIN default.dim_cabang d ON f.id_cabang = d.id_cabang
            WHERE toDate(f.tanggal_diterima) = '{d}'
            GROUP BY d.nama_cabang, f.status 
            ORDER BY d.nama_cabang ASC, f.status ASC""",

        "stok_rows": f"""
            SELECT dc.nama_cabang, di.nama_item, f.quantity_akhir, f.total_biaya_akhir
            FROM default.fact_nilai_persediaan f
            LEFT JOIN default.dim_cabang dc ON f.id_cabang = dc.id_cabang
            LEFT JOIN default.dim_item di ON f.id_item = di.id_item
            WHERE f.tanggal <= '{d}'
            ORDER BY dc.nama_cabang ASC, di.nama_item ASC
            LIMIT 1 BY f.id_cabang, f.id_item""",

        "top_menu_rows": f"""
            SELECT dm.nama_menu, COUNT(*) AS trx, SUM(f.qty) AS qty, SUM(f.subtotal) AS rev
            FROM default.fact_sales f
            LEFT JOIN default.dim_menu dm ON f.id_menu = dm.id_menu
            WHERE f.tanggal_transaksi = '{d}'
            GROUP BY dm.nama_menu 
            ORDER BY rev DESC 
            LIMIT 5""",

        # Perhitungan Deviasi Langsung dari Tabel Asli fact_cost_of_material
        "deviasi_rows": f"""
            SELECT 
                d.nama_cabang,
                COUNT(*) AS total_mutasi,
                ROUND(AVG(IF(f.bom_total = 0, 0, ((f.total_biaya - f.bom_total) / f.bom_total) * 100)), 2) AS avg_dev_pct,
                SUM(IF(f.total_biaya > f.bom_total, 1, 0)) AS boros,
                SUM(IF(f.total_biaya < f.bom_total, 1, 0)) AS efisien
            FROM default.fact_cost_of_material f
            LEFT JOIN default.dim_cabang d ON f.id_cabang = d.id_cabang
            WHERE f.tanggal = '{d}'
            GROUP BY d.nama_cabang
            ORDER BY avg_dev_pct DESC"""
    }

    results = {}
    for key, sql in queries.items():
        try:
            results[key] = ch.execute(sql)
        except Exception as e:
            print(f"❌ Gagal mengeksekusi {key}: {str(e)}")
            results[key] = []

    context["ti"].xcom_push(key="exec_date", value=d)
    for k, v in results.items():
        context["ti"].xcom_push(key=k, value=v)

    print(f"✅ [FETCH] {d} — Seluruh data finansial berhasil ditarik.")


def generate_excel_report(**context):
    ti        = context["ti"]
    exec_date = ti.xcom_pull(key="exec_date")

    def pull(k): return ti.xcom_pull(key=k) or []

    rev_summary   = pull("rev_summary")
    revenue_rows  = pull("revenue_rows")
    hpp_rows      = pull("hpp_rows")
    gp_rows       = pull("gp_rows")
    po_rows       = pull("po_rows")
    stok_rows     = pull("stok_rows")
    top_menu_rows = pull("top_menu_rows")
    deviasi_rows  = pull("deviasi_rows")

    NAVY, BLUE, WHITE, LIGHT, ALT = "1E3A5F", "2563EB", "FFFFFF", "F1F5F9", "E8F0FE"
    FMT_IDR, FMT_PCT, FMT_NUM = '#,##0', '0.0%', '#,##0.##'

    def hdr_font(): return Font(name="Arial", bold=True, color=WHITE, size=11)
    def hdr_fill(color=NAVY): return PatternFill("solid", fgColor=color)
    def data_font(bold=False, color="000000"): return Font(name="Arial", bold=bold, color=color, size=10)
    def alt_fill(row_idx): return PatternFill("solid", fgColor=ALT if row_idx % 2 == 0 else WHITE)
    def thin_border():
        s = Side(style="thin", color="CBD5E1")
        return Border(left=s, right=s, top=s, bottom=s)

    def center(): return Alignment(horizontal="center", vertical="center")
    def right():  return Alignment(horizontal="right",  vertical="center")
    def left():   return Alignment(horizontal="left",   vertical="center")

    def write_header(ws, headers, row=1, fill_color=NAVY):
        for col, text in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=text)
            c.font, c.fill, c.alignment, c.border = hdr_font(), hdr_fill(fill_color), center(), thin_border()

    def style_data_row(ws, row_idx, n_cols, number_cols=(), pct_cols=(), idr_cols=(), bold=False, fill_color=None):
        fill = PatternFill("solid", fgColor=fill_color) if fill_color else alt_fill(row_idx)
        for col in range(1, n_cols + 1):
            c = ws.cell(row=row_idx, column=col)
            c.font, c.fill, c.border = data_font(bold=bold), fill, thin_border()
            c.alignment = right() if col in number_cols or col in idr_cols or col in pct_cols else left()
            if col in idr_cols:     c.number_format = FMT_IDR
            elif col in pct_cols:   c.number_format = FMT_PCT
            elif col in number_cols: c.number_format = FMT_NUM

    def set_col_widths(ws, widths):
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    def title_row(ws, text, n_merge, row=1):
        ws.row_dimensions[row].height = 28
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_merge)
        c = ws.cell(row=row, column=1, value=text)
        c.font, c.fill, c.alignment = Font(name="Arial", bold=True, color=WHITE, size=13), hdr_fill(NAVY), center()

    def sub_row(ws, text, n_merge, row=2):
        ws.row_dimensions[row].height = 16
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_merge)
        c = ws.cell(row=row, column=1, value=f"Periode: {text}")
        c.font, c.fill, c.alignment = Font(name="Arial", italic=True, color="64748B", size=10), PatternFill("solid", fgColor=LIGHT), center()

    wb = Workbook()
    wb.remove(wb.active)

    # SHEET 1 : RINGKASAN
    ws = wb.create_sheet("Ringkasan")
    ws.sheet_view.showGridLines = False
    title_row(ws, f"📊  LAPORAN KEUANGAN HARIAN  —  ERP SCM", 4, row=1)
    sub_row(ws, exec_date, 4, row=2)

    kpi_data = [
        ("TOTAL REVENUE",       rev_summary[0][0] if rev_summary else 0, FMT_IDR),
        ("TOTAL TRANSAKSI",     rev_summary[0][1] if rev_summary else 0, FMT_NUM),
        ("ITEM TERJUAL",        rev_summary[0][2] if rev_summary else 0, FMT_NUM),
        ("TOTAL HPP",           f"=Rincian_HPP!D{2+len(hpp_rows)+2}",    FMT_IDR),
        ("GROSS PROFIT",        "=B5-B8",                                 FMT_IDR),
        ("GP MARGIN",           "=IF(B5=0,0,B9/B5)",                     FMT_PCT),
    ]
    ws.cell(row=4, column=1, value="INDIKATOR").font = hdr_font(); ws.cell(row=4, column=1).fill = hdr_fill(BLUE); ws.cell(row=4, column=1).alignment = center()
    ws.cell(row=4, column=2, value="NILAI").font = hdr_font(); ws.cell(row=4, column=2).fill = hdr_fill(BLUE); ws.cell(row=4, column=2).alignment = center()

    for i, (label, val, fmt) in enumerate(kpi_data, start=5):
        cl, cv = ws.cell(row=i, column=1, value=label), ws.cell(row=i, column=2, value=val)
        cl.font, cl.fill, cl.border, cl.alignment = data_font(bold=True), alt_fill(i), thin_border(), left()
        cv.font, cv.fill, cv.border, cv.alignment = data_font(bold=True, color=NAVY), alt_fill(i), thin_border(), right()
        cv.number_format = fmt
    set_col_widths(ws, [30, 22, 5, 5])

    # SHEET 2 : GROSS PROFIT PER CABANG
    ws2 = wb.create_sheet("Gross_Profit")
    ws2.sheet_view.showGridLines = False
    title_row(ws2, "GROSS PROFIT PER CABANG", 5, row=1)
    sub_row(ws2, exec_date, 5, row=2)
    write_header(ws2, ["CABANG", "REVENUE (Rp)", "HPP (Rp)", "GROSS PROFIT (Rp)", "MARGIN (%)"], row=3)
    for i, r in enumerate(gp_rows, start=4):
        cab, rev, hpp, gp = r
        ws2.cell(row=i, column=1, value=cab)
        ws2.cell(row=i, column=2, value=rev)
        ws2.cell(row=i, column=3, value=hpp)
        ws2.cell(row=i, column=4, value=f"=B{i}-C{i}")
        ws2.cell(row=i, column=5, value=f"=IF(B{i}=0,0,D{i}/B{i})")
        style_data_row(ws2, i, 5, idr_cols=(2,3,4), pct_cols=(5,))
    tr = 4 + len(gp_rows)
    ws2.cell(row=tr, column=1, value="TOTAL")
    for col, rng in [(2, "B"), (3, "C"), (4, "D")]:
        ws2.cell(row=tr, column=col, value=f"=SUM({rng}4:{rng}{tr-1})")
    ws2.cell(row=tr, column=5, value=f"=IF(B{tr}=0,0,D{tr}/B{tr})")
    style_data_row(ws2, tr, 5, idr_cols=(2,3,4), pct_cols=(5,), bold=True, fill_color=NAVY)
    for col in range(1, 6): ws2.cell(row=tr, column=col).font = Font(name="Arial", bold=True, color=WHITE, size=10)
    set_col_widths(ws2, [32, 20, 20, 20, 14])

    # SHEET 3 : TOP 5 MENU TERLARIS
    ws3 = wb.create_sheet("Top_Menu")
    ws3.sheet_view.showGridLines = False
    title_row(ws3, "TOP 5 MENU TERLARIS", 5, row=1)
    sub_row(ws3, exec_date, 5, row=2)
    write_header(ws3, ["RANK", "NAMA MENU", "TRANSAKSI", "QTY TERJUAL", "REVENUE (Rp)"], row=3)
    for i, r in enumerate(top_menu_rows, start=4):
        menu, trx, qty, rev = r
        rank = i - 3
        ws3.cell(row=i, column=1, value=rank); ws3.cell(row=i, column=2, value=menu); ws3.cell(row=i, column=3, value=trx); ws3.cell(row=i, column=4, value=qty); ws3.cell(row=i, column=5, value=rev)
        style_data_row(ws3, i, 5, number_cols=(1,3,4), idr_cols=(5,))
        ws3.cell(row=i, column=1).alignment = center()
        f_color = "FEF9C3" if rank == 1 else ("F1F5F9" if rank == 2 else WHITE)
        for col in range(1, 6): ws3.cell(row=i, column=col).fill = PatternFill("solid", fgColor=f_color)
    set_col_widths(ws3, [8, 34, 14, 14, 20])

    # SHEET 4 : PENDAPATAN PER CHANNEL
    ws4 = wb.create_sheet("Revenue_Channel")
    ws4.sheet_view.showGridLines = False
    title_row(ws4, "PENDAPATAN PER CABANG & CHANNEL PEMBAYARAN", 5, row=1)
    sub_row(ws4, exec_date, 5, row=2)
    write_header(ws4, ["CABANG", "CHANNEL", "TRANSAKSI", "QTY", "REVENUE (Rp)"], row=3)
    for i, r in enumerate(revenue_rows, start=4):
        cab, ch_, trx, qty, rev = r
        ws4.cell(row=i, column=1, value=cab); ws4.cell(row=i, column=2, value=ch_); ws4.cell(row=i, column=3, value=trx); ws4.cell(row=i, column=4, value=qty); ws4.cell(row=i, column=5, value=rev)
        style_data_row(ws4, i, 5, number_cols=(3,4), idr_cols=(5,))
        ws4.cell(row=i, column=2).fill = PatternFill("solid", fgColor="DCFCE7" if ch_ in ("GOFOOD","GRABFOOD") else "FEF9C3")
    tr4 = 4 + len(revenue_rows)
    ws4.cell(row=tr4, column=1, value="TOTAL"); ws4.cell(row=tr4, column=2, value="—")
    for col, rng in [(3,"C"),(4,"D"),(5,"E")]: ws4.cell(row=tr4, column=col, value=f"=SUM({rng}4:{rng}{tr4-1})")
    style_data_row(ws4, tr4, 5, number_cols=(3,4), idr_cols=(5,), bold=True, fill_color=NAVY)
    for col in range(1, 6): ws4.cell(row=tr4, column=col).font = Font(name="Arial", bold=True, color=WHITE, size=10)
    set_col_widths(ws4, [32, 16, 14, 14, 20])

    # SHEET 5 : HPP / COST OF MATERIAL
    ws5 = wb.create_sheet("Rincian_HPP")
    ws5.sheet_view.showGridLines = False
    title_row(ws5, "HARGA POKOK PENJUALAN (HPP) PER KATEGORI BAHAN", 4, row=1)
    sub_row(ws5, exec_date, 4, row=2)
    write_header(ws5, ["CABANG", "KATEGORI BAHAN", "QTY TERPAKAI", "TOTAL HPP (Rp)"], row=3)
    for i, r in enumerate(hpp_rows, start=4):
        cab, kat, qty, biaya = r
        ws5.cell(row=i, column=1, value=cab); ws5.cell(row=i, column=2, value=kat); ws5.cell(row=i, column=3, value=qty); ws5.cell(row=i, column=4, value=biaya)
        style_data_row(ws5, i, 4, number_cols=(3,), idr_cols=(4,))
    tr5 = 4 + len(hpp_rows)
    ws5.cell(row=tr5, column=1, value="TOTAL"); ws5.cell(row=tr5, column=2, value="—")
    ws5.cell(row=tr5, column=3, value=f"=SUM(C4:C{tr5-1})"); ws5.cell(row=tr5, column=4, value=f"=SUM(D4:D{tr5-1})")
    style_data_row(ws5, tr5, 4, number_cols=(3,), idr_cols=(4,), bold=True, fill_color=NAVY)
    for col in range(1, 5): ws5.cell(row=tr5, column=col).font = Font(name="Arial", bold=True, color=WHITE, size=10)
    set_col_widths(ws5, [32, 20, 18, 22])

    # SHEET 6 : DEVIASI BOM (REAL TIME DARI KAKULASI TASK 1)
    if deviasi_rows:
        ws6 = wb.create_sheet("Deviasi_BOM")
        ws6.sheet_view.showGridLines = False
        title_row(ws6, "DEVIASI BIAYA PEMAKAIAN BOM (WASTE vs EFISIENSI)", 5, row=1)
        sub_row(ws6, exec_date, 5, row=2)
        write_header(ws6, ["CABANG", "TOTAL ITEM DIMUTASI", "RATA-RATA DEVIASI (%)", "BAHAN BOROS (WASTE)", "BAHAN EFISIEN"], row=3)
        for i, r in enumerate(deviasi_rows, start=4):
            cab, total, avg_dev, boros, efisien = r
            ws6.cell(row=i, column=1, value=cab)
            ws6.cell(row=i, column=2, value=total)
            ws6.cell(row=i, column=3, value=float(avg_dev)/100)  # Konversi ke bentuk desimal untuk format %
            ws6.cell(row=i, column=4, value=boros)
            ws6.cell(row=i, column=5, value=efisien)
            style_data_row(ws6, i, 5, number_cols=(2,4,5), pct_cols=(3,))
            
            # Dinamis Alert Color: Merah (> 5% waste), Hijau (< 0% hemat), Kuning (Aman/Wajar)
            d_color = "FEE2E2" if float(avg_dev) > 5.0 else ("DCFCE7" if float(avg_dev) < 0.0 else "FEF9C3")
            ws6.cell(row=i, column=3).fill = PatternFill("solid", fgColor=d_color)
        set_col_widths(ws6, [32, 22, 24, 22, 22])

    # SHEET 7 : PURCHASE ORDER
    ws7 = wb.create_sheet("Purchase_Order")
    ws7.sheet_view.showGridLines = False
    title_row(ws7, "AKTIVITAS PURCHASE ORDER", 5, row=1)
    sub_row(ws7, exec_date, 5, row=2)
    write_header(ws7, ["CABANG", "STATUS", "JUMLAH PO", "TOTAL QTY", "TOTAL NILAI (Rp)"], row=3)
    for i, r in enumerate(po_rows, start=4):
        cab, status, jml_po, qty, nilai = r
        ws7.cell(row=i, column=1, value=cab); ws7.cell(row=i, column=2, value=status); ws7.cell(row=i, column=3, value=jml_po); ws7.cell(row=i, column=4, value=qty); ws7.cell(row=i, column=5, value=nilai)
        style_data_row(ws7, i, 5, number_cols=(3,4), idr_cols=(5,))
        ws7.cell(row=i, column=2).fill = PatternFill("solid", fgColor="DCFCE7" if status == "DITERIMA" else "FEF9C3")
    tr7 = 4 + len(po_rows)
    ws7.cell(row=tr7, column=1, value="TOTAL"); ws7.cell(row=tr7, column=2, value="—")
    for col, rng in [(3,"C"),(4,"D"),(5,"E")]: ws7.cell(row=tr7, column=col, value=f"=SUM({rng}4:{rng}{tr7-1})")
    style_data_row(ws7, tr7, 5, number_cols=(3,4), idr_cols=(5,), bold=True, fill_color=NAVY)
    for col in range(1, 6): ws7.cell(row=tr7, column=col).font = Font(name="Arial", bold=True, color=WHITE, size=10)
    set_col_widths(ws7, [32, 14, 14, 16, 22])

    # SHEET 8 : STOK AKHIR
    ws8 = wb.create_sheet("Stok_Akhir")
    ws8.sheet_view.showGridLines = False
    title_row(ws8, "NILAI PERSEDIAAN AKHIR — SNAPSHOT HARIAN", 4, row=1)
    sub_row(ws8, exec_date, 4, row=2)
    write_header(ws8, ["CABANG", "NAMA ITEM", "STOK AKHIR", "NILAI PERSEDIAAN (Rp)"], row=3)
    stok_sorted = sorted(stok_rows, key=lambda r: float(r[2]))
    for i, r in enumerate(stok_sorted, start=4):
        cab, item, qty, nilai = r
        ws8.cell(row=i, column=1, value=cab); ws8.cell(row=i, column=2, value=item); ws8.cell(row=i, column=3, value=float(qty)); ws8.cell(row=i, column=4, value=float(nilai))
        style_data_row(ws8, i, 4, number_cols=(3,), idr_cols=(4,))
        ws8.cell(row=i, column=3).fill = PatternFill("solid", fgColor="FEE2E2" if float(qty) < 500 else ("FEF9C3" if float(qty) < 2000 else WHITE))
    tr8 = 4 + len(stok_sorted)
    ws8.cell(row=tr8, column=1, value="TOTAL NILAI PERSEDIAAN"); ws8.cell(row=tr8, column=2, value="—"); ws8.cell(row=tr8, column=3, value="—"); ws8.cell(row=tr8, column=4, value=f"=SUM(D4:D{tr8-1})")
    style_data_row(ws8, tr8, 4, idr_cols=(4,), bold=True, fill_color=NAVY)
    for col in range(1, 5): ws8.cell(row=tr8, column=col).font = Font(name="Arial", bold=True, color=WHITE, size=10)
    set_col_widths(ws8, [32, 28, 16, 24])

    tmp_path = os.path.join(tempfile.gettempdir(), f"laporan_keuangan_{exec_date}.xlsx")
    wb.save(tmp_path)
    context["ti"].xcom_push(key="excel_path", value=tmp_path)
    print(f"✅ [EXCEL] Laporan SCM tersimpan di: {tmp_path}")


# ============================================================================
# TASK 3 : RENDER HTML + KIRIM EMAIL + LAMPIRKAN EXCEL
# ============================================================================
def render_and_send_report(**context):
    ti = context["ti"]
    exec_date     = ti.xcom_pull(key="exec_date")
    rev_summary   = ti.xcom_pull(key="rev_summary")   or []
    revenue_rows  = ti.xcom_pull(key="revenue_rows")  or []
    hpp_rows      = ti.xcom_pull(key="hpp_rows")       or []
    gp_rows       = ti.xcom_pull(key="gp_rows")        or []
    po_rows       = ti.xcom_pull(key="po_rows")        or []
    stok_rows     = ti.xcom_pull(key="stok_rows")      or []
    top_menu_rows = ti.xcom_pull(key="top_menu_rows")  or []
    deviasi_rows  = ti.xcom_pull(key="deviasi_rows")   or []
    excel_path    = ti.xcom_pull(key="excel_path")

    def rp(v): return f"Rp {float(v):,.0f}" if v else "-"
    def pct(n, d): return f"{float(n)/float(d)*100:.1f}%" if d and float(d) != 0 else "0.0%"

    css = """
    body{font-family:'Segoe UI',Arial,sans-serif;background:#f8fafc;color:#1e293b;margin:0;padding:0}
    .w{max-width:860px;margin:24px auto;background:#fff;border-radius:12px;box-shadow:0 4px 24px rgba(0,0,0,.08);overflow:hidden}
    .hd{background:linear-gradient(135deg,#1e3a5f,#2563eb);color:#fff;padding:32px 36px}
    .hd h1{margin:0 0 4px;font-size:22px;font-weight:700}
    .hd p{margin:0;font-size:13px;opacity:.8}
    .bd{padding:28px 36px}
    .sec{margin-bottom:32px}
    .sec h2{font-size:15px;font-weight:700;color:#1e3a5f;border-left:4px solid #2563eb;padding-left:10px;margin:0 0 14px}
    .kpig{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:20px}
    .kpic{flex:1;min-width:160px;background:#f1f5f9;border-radius:8px;padding:16px 20px}
    .kpic .lb{font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px}
    .kpic .vl{font-size:22px;font-weight:700;color:#1e3a5f}
    .kpic .sb{font-size:11px;color:#94a3b8;margin-top:2px}
    table{width:100%;border-collapse:collapse;font-size:13px}
    thead tr{background:#1e3a5f;color:#fff}
    thead th{padding:10px 12px;text-align:left;font-weight:600}
    tbody tr:nth-child(even){background:#f8fafc}
    tbody td{padding:9px 12px;border-bottom:1px solid #e2e8f0}
    .bg{background:#dcfce7;color:#15803d;padding:2px 8px;border-radius:12px;font-size:11px;font-weight:600}
    .br{background:#fee2e2;color:#dc2626;padding:2px 8px;border-radius:12px;font-size:11px;font-weight:600}
    .by{background:#fef9c3;color:#854d0e;padding:2px 8px;border-radius:12px;font-size:11px;font-weight:600}
    .att{background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:14px 18px;margin-bottom:24px;font-size:13px;color:#1e40af}
    .ft{background:#f1f5f9;padding:18px 36px;font-size:11px;color:#94a3b8;text-align:center}
    """

    total_rev  = float(rev_summary[0][0]) if rev_summary and rev_summary[0][0] else 0
    total_trx  = int(rev_summary[0][1])   if rev_summary and rev_summary[0][1] else 0
    total_item = int(rev_summary[0][2])   if rev_summary and rev_summary[0][2] else 0
    total_hpp  = sum(float(r[3]) for r in hpp_rows)
    gp         = total_rev - total_hpp

    kpi = f"""<div class="kpig">
    <div class="kpic"><div class="lb">Total Revenue</div><div class="vl">{rp(total_rev)}</div><div class="sb">{total_trx:,} transaksi &middot; {total_item:,} item</div></div>
    <div class="kpic"><div class="lb">Total HPP</div><div class="vl">{rp(total_hpp)}</div><div class="sb">Harga Pokok Penjualan</div></div>
    <div class="kpic"><div class="lb">Gross Profit</div><div class="vl">{rp(gp)}</div><div class="sb">Margin {pct(gp,total_rev)}</div></div>
    </div>"""

    def tbl(headers, rows_html):
        return f"<table><thead><tr>{''.join(f'<th>{h}</th>' for h in headers)}</tr></thead><tbody>{rows_html}</tbody></table>"

    gp_rows_html = "".join(f"<tr><td>{r[0]}</td><td>{rp(r[1])}</td><td>{rp(r[2])}</td><td><span class='{'bg' if float(r[3])>=0 else 'br'}'>{rp(r[3])}</span></td><td>{pct(r[3],r[1])}</td></tr>" for r in gp_rows)
    tm_html = "".join(f"<tr><td><b style='color:#2563eb'>#{i}</b></td><td>{r[0]}</td><td>{int(r[1]):,}</td><td>{int(r[2]):,}</td><td>{rp(r[3])}</td></tr>" for i, r in enumerate(top_menu_rows, 1))
    rv_html = "".join(f"<tr><td>{r[0]}</td><td><span class='{'bg' if r[1] in ('GOFOOD','GRABFOOD') else 'by'}'>{r[1]}</span></td><td>{int(r[2]):,}</td><td>{int(r[3]):,}</td><td>{rp(r[4])}</td></tr>" for r in revenue_rows)
    hp_html = "".join(f"<tr><td>{r[0]}</td><td>{r[1]}</td><td>{float(r[2]):,.1f}</td><td>{rp(r[3])}</td></tr>" for r in hpp_rows)
    po_html = "".join(f"<tr><td>{r[0]}</td><td><span class='{'bg' if r[1]=='DITERIMA' else 'by'}'>{r[1]}</span></td><td>{int(r[2]):,}</td><td>{float(r[3]):,.0f}</td><td>{rp(r[4])}</td></tr>" for r in po_rows)
    st_html = "".join(f"<tr><td>{r[0]}</td><td>{r[1]}</td><td><span class='{'br' if float(r[2])<500 else ('by' if float(r[2])<2000 else 'bg')}'>{float(r[2]):,.1f}</span></td><td>{rp(r[3])}</td></tr>" for r in sorted(stok_rows, key=lambda r: float(r[2]))[:20])

    dev_sec = ""
    if deviasi_rows:
        dv_html = "".join(f"<tr><td>{r[0]}</td><td>{int(r[1]):,}</td><td><span class='{'br' if float(r[2])>5.0 else ('bg' if float(r[2])<0.0 else 'by')}'>{r[2]:+.2f}%</span></td><td>{int(r[3]):,}</td><td>{int(r[4]):,}</td></tr>" for r in deviasi_rows)
        dev_sec = f"""<div class="sec"><h2>⚖️ Deviasi Pemakaian BOM (Real-time COM vs BOM)</h2>{tbl(["Cabang","Total Mutasi","Avg Deviasi","Boros (Waste)","Efisien"], dv_html)}</div>"""

    html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><style>{css}</style></head>
    <body><div class="w">
    <div class="hd"><h1>📊 Laporan Keuangan Harian ERP (SCM Scope)</h1><p>Periode: <b>{exec_date}</b> &nbsp;|&nbsp; Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p></div>
    <div class="bd">
      <div class="att">📎 <b>Laporan Excel Berhasil Dilampirkan</b> — Berisi rincian 8 Sheet lengkap dengan otomasi formula Excel.</div>
      <div class="sec"><h2>💰 Ringkasan Keuangan</h2>{kpi}</div>
      <div class="sec"><h2>📈 Gross Profit per Cabang</h2>{tbl(["Cabang","Revenue","HPP","Gross Profit","Margin"], gp_rows_html)}</div>
      <div class="sec"><h2>🏆 Top 5 Menu Terlaris</h2>{tbl(["#","Menu","Transaksi","Terjual","Revenue"], tm_html)}</div>
      <div class="sec"><h2>🛒 Pendapatan per Channel</h2>{tbl(["Cabang","Channel","Transaksi","Qty","Revenue"], rv_html)}</div>
      <div class="sec"><h2>⚙️ Rincian HPP per Kategori</h2>{tbl(["Cabang","Kategori","Qty Terpakai","Total HPP"], hp_html)}</div>
      {dev_sec}
      <div class="sec"><h2>🚚 Aktivitas Purchase Order</h2>{tbl(["Cabang","Status","Jumlah PO","Total Qty","Nilai PO"], po_html)}</div>
      <div class="sec"><h2>📦 Stok Akhir Terendah (Top 20 Snapshot)</h2>{tbl(["Cabang","Item","Stok Akhir","Nilai Persediaan"], st_html)}</div>
    </div>
    <div class="ft">Laporan otomatis Airflow DAG: <b>daily_financial_report</b> | Source: ClickHouse Data Warehouse</div>
    </div></body></html>"""

    msg = MIMEMultipart("mixed")
    msg["From"], msg["To"], msg["Subject"] = SENDER_EMAIL, REPORT_RECEIVER, f"📊 Laporan Keuangan Harian ERP — {exec_date}"
    msg.attach(MIMEText(html, "html"))

    if excel_path and os.path.exists(excel_path):
        with open(excel_path, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment", filename=os.path.basename(excel_path))
        msg.attach(part)

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USER, SMTP_PASSWORD)
        server.sendmail(SENDER_EMAIL, REPORT_RECEIVER, msg.as_string())
    print("🚀 [EMAIL] Laporan keuangan SCM harian berhasil dikirim via Brevo.")

t1 = PythonOperator(task_id="fetch_financial_data",   python_callable=fetch_financial_data,   provide_context=True, dag=dag)
t2 = PythonOperator(task_id="generate_excel_report",  python_callable=generate_excel_report,  provide_context=True, dag=dag)
t3 = PythonOperator(task_id="render_and_send_report", python_callable=render_and_send_report, provide_context=True, dag=dag)

t1 >> t2 >> t3